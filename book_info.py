import openpyxl

book_data = {
        'information_book1': {'978-600-353-265-6': {'Title': "Sense and Sensibility", 'Author': 'Jane Austen', 'Genre': 'English Novels', 'Edition': '5', 'Date': '2019'}},
        'information_book2': {'978-964-431-100-0': {'Title': 'The Little Prince', 'Author': 'Antoine de Saint-Exupery', 'Genre': 'French Novels', 'Edition': '3', 'Date': '2019'}},
        'information_book3': {'978-964-7518-12-3': {'Title': 'Madly', 'Author': 'Christiane Rochefort', 'Genre': 'French Novels', 'Edition': '10', 'Date': '2019'}},
        'information_book4': {'978-600-08-0862-4': {'Title': "Emma", 'Author': 'Jane Austen', 'Genre': 'English Novels', 'Edition': '1', 'Date': '2020'}},
        'information_book5': {'978-600-8866-32-9': {'Title': "Stupidity", 'Author': 'Xavier Clement', 'Genre': 'Self-Knowledge', 'Edition': '7', 'Date': '2021'}},
        'information_book6': {'978-600-8852-21-6': {'Title': "Room", 'Author': 'Emma Donoghue', 'Genre': 'English Novels', 'Edition': '1', 'Date': '2017'}},
        'information_book7': {'978-600-5364-13-2': {'Title': "Four Works by Florence", 'Author': 'Florence Scovel Shinn', 'Genre': 'New Thought', 'Edition': '9', 'Date': '2009'}},
        'information_book8': {'978-964-95477-1-1': {'Title': "The Blind Owl", 'Author': 'Sadegh Hedayat', 'Genre': 'Persian Novels', 'Edition': '1', 'Date': '2004'}},
        'information_book9': {'978-600-8951-05-6': {'Title': "The Shadow Effect", 'Author': 'Debbie Ford', 'Genre': 'Self-Help', 'Edition': '3', 'Date': '2020'}},
        'information_book10': {'978-964-6516-43-2': {'Title': "The Dark Side of the Light Chasers", 'Author': 'Debbie Ford', 'Genre': 'Success', 'Edition': '10', 'Date': '2008'}},
        'information_book11': {'978-600-384-033-1': {'Title': "The Couple Next Door", 'Author': 'Shari Lapena', 'Genre': 'Canadian Novels', 'Edition': '5', 'Date': '2018'}},
        'information_book12': {'978-964-311-919-5': {'Title': "The Bastard of Istanbul", 'Author': 'Elif Shafak', 'Genre': 'Turkish Novels', 'Edition': '1', 'Date': '2015'}},
        'information_book13': {'978-964-8175-01-2': {'Title': "The Other Father", 'Author': 'Parnioush Saniee', 'Genre': 'Persian Novels', 'Edition': '8', 'Date': '2009'}},
        'information_book14': {'978-622-201-946-4': {'Title': "The Confessions of Harry Leder", 'Author': 'Walter Tevis', 'Genre': 'American Novels', 'Edition': '1', 'Date': '2021'}},
        'information_book15': {'7': {'Title': "Animal Farm", 'Author': 'George Orwell', 'Genre': 'English Novels', 'Edition': '5', 'Date': '2021'}}
}


class BookData:
    def __init__(self, book_data):
        self.book_data = book_data

    def add_book(self, isbn, title, author, genre, edition, date):

        book_info = {
            'Title': title,
            'Author': author,
            'Genre': genre,
            'Edition': edition,
            'Date': date
        }

        for book_key, book_dict in self.book_data.items():
            for isbn2, book_details in book_dict.items():
                if isbn == isbn2:
                    return False

        book_key = f"information_book{len(self.book_data) + 1}"
        self.book_data[book_key] = {isbn: book_info}
        return True

    def delete_book(self, filename, identifier):
        workbook = openpyxl.load_workbook(filename)
        identifier = str(identifier).lower()

        for worksheet in workbook.worksheets:
            isbn = str(worksheet.cell(row=4, column=5).value).lower() if worksheet.cell(row=4, column=5).value else None
            title = str(worksheet.cell(row=5, column=5).value).lower() if worksheet.cell(row=5, column=5).value else None

            if identifier == isbn or identifier == title:
                for key, value in self.book_data.items():
                    if isbn in value:
                        del self.book_data[key]
                        break

                workbook.remove(worksheet)
                workbook.save(filename)
                return True

        return False


    def show_books(self, filename):
        workbook = openpyxl.load_workbook(filename)

        worksheet_names = workbook.sheetnames

        worksheet_names.sort()

        data = []
        for worksheet_name in worksheet_names:
            worksheet = workbook[worksheet_name]

            if any(worksheet.values):

                for row in worksheet.iter_rows(min_row=1, min_col=1, values_only=True):

                    row_data = [value for value in row if value is not None and value != []]

                    if row_data:
                        data.append(row_data)

                data.append([])  

        return data
    
    def search_book(self, identifier):
        identifier = str(identifier).lower()
        book_info = None

        for book_key, book_dict in self.book_data.items():
            for isbn, book_details in book_dict.items():
                if identifier == isbn.lower() or identifier == book_details['Title'].lower():
                    book_info = book_details
                    break

        return book_info


book_info = BookData(book_data)